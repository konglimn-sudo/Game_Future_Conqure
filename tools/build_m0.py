# 生成 M0 数值原型工作簿：参数 + 4 情景 × 20 回合 + 对比
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.utils import get_column_letter

ARIAL = "Arial"
BLUE = Font(name=ARIAL, color="0000FF")          # 可改输入
BLACK = Font(name=ARIAL)
GREEN = Font(name=ARIAL, color="008000")         # 跨表引用
BOLD = Font(name=ARIAL, bold=True)
HDR_FILL = PatternFill("solid", start_color="1F4E79")
HDR_FONT = Font(name=ARIAL, bold=True, color="FFFFFF", size=9)
WRAP_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Border(bottom=Side(style="thin", color="D9D9D9"))

wb = Workbook()

# ---------------- 说明 ----------------
ws = wb.active
ws.title = "说明"
ws.column_dimensions["A"].width = 100
lines = [
    ("《Future Conquer》M0 数值原型", 16, True),
    ("", 11, False),
    ("目的：验证设计文档第 2 节的核心机制——算力三通道分配（训练/推理/研发）的取舍是否『有感觉』。", 11, False),
    ("模型为单势力自循环，刻意不含战斗、外交、轨道层（那些属于 M2+）。", 11, False),
    ("", 11, False),
    ("怎么用：", 12, True),
    ("· 蓝色单元格 = 可修改的输入：『参数』表全部数值；各情景表的 训练%/推理%（研发% 自动 = 1 - 两者，逐回合可调）。", 11, False),
    ("· 修改后用 Excel/Numbers/LibreOffice 打开会自动重算；『对比』表汇总四种策略的关键指标与走势图。", 11, False),
    ("", 11, False),
    ("模型简化假设：", 12, True),
    ("· 芯片每回合自动投资数据中心扩容（受投资上限约束），扩容次回合生效。", 11, False),
    ("· 推理需求随版图扩张线性增长；推理供给率不足会压低经济系数，拖累芯片/数据产出（滞后一回合）。", 11, False),
    ("· 科技等级提升芯片与数据产出；代际提升算力效率系数与军力。", 11, False),
    ("", 11, False),
    ("验收观察点（M0 通过标准）：", 12, True),
    ("1. 电力墙：中后期『⚡受限』出现，算力被电力供给锁死 → 验证『电力是上限约束』。", 11, False),
    ("2. 数据弹药：奇点冲刺流的训练数据满足率跌破 100% → 验证『数据是训练瓶颈』。", 11, False),
    ("3. 推理饥饿：奇点流经济系数走低，产能被拖累 → 抽算力冲代际的代价当回合可见。", 11, False),
    ("4. 军备 vs 代际：军备流军力先发优势明显；20 回合内奇点流军力难以反超 → 印证奇点路线需要外交/隐蔽保护期（设计文档 6.3、7 章）。", 11, False),
    ("", 11, False),
    ("若以上四条都能在数值上看到，且改动参数能明显改变格局，M0 验收通过，进入 M1（Godot 经济沙盒）。", 11, False),
]
for i, (txt, size, bold) in enumerate(lines, 1):
    c = ws.cell(row=i, column=1, value=txt)
    c.font = Font(name=ARIAL, size=size, bold=bold)
    c.alignment = Alignment(wrap_text=True, vertical="top")

# ---------------- 参数 ----------------
ps = wb.create_sheet("参数")
params = [
    ("初始电力供给", 100, "电力/回合，决定算力上限"),
    ("电力每回合增长", 6, "电网建设速度"),
    ("每点算力耗电", 1, "电力→算力转换"),
    ("初始数据中心容量", 60, "算力点"),
    ("新建容量芯片成本", 2, "芯片/算力点"),
    ("每回合芯片投资上限", 20, "建造产能瓶颈"),
    ("初始芯片库存", 30, ""),
    ("晶圆厂基础芯片产出", 10, "芯片/回合"),
    ("区域数", 6, "版图规模"),
    ("每区域数据产出", 12, "数据/回合"),
    ("初始数据池", 60, ""),
    ("训练每算力耗数据", 0.8, "数据是训练弹药"),
    ("Gen2 门槛（累积训练）", 400, ""),
    ("Gen3 门槛", 900, ""),
    ("Gen4 门槛", 1800, ""),
    ("每代际能力系数增量", 0.4, "代际→算力效率与军力乘数"),
    ("科技每级门槛（科技点）", 80, ""),
    ("科技每级产出加成", 0.05, "作用于芯片与数据产出"),
    ("推理基础需求", 30, "维持现有版图与部队"),
    ("推理需求每回合增长", 3, "版图扩张带来的需求"),
    ("经济系数底线", 0.5, "推理完全断供时产出打五折"),
]
for col, w in (("A", 26), ("B", 10), ("C", 34)):
    ps.column_dimensions[col].width = w
for i, h in enumerate(["参数", "值", "说明"], 1):
    c = ps.cell(row=1, column=i, value=h)
    c.font, c.fill, c.alignment = HDR_FONT, HDR_FILL, WRAP_CENTER
for r, (name, val, note) in enumerate(params, 2):
    ps.cell(row=r, column=1, value=name).font = BLACK
    v = ps.cell(row=r, column=2, value=val)
    v.font = BLUE
    v.number_format = "0%" if name == "科技每级产出加成" else ("0.0" if isinstance(val, float) else "0")
    ps.cell(row=r, column=3, value=note).font = Font(name=ARIAL, size=9, color="808080")

P = lambda r: f"参数!$B${r}"  # 参数行号: 2..22 与 params 顺序对应
(P_POW0, P_POWG, P_E_PER, P_DC0, P_CAPCOST, P_INVCAP, P_CHIP0, P_FAB, P_REG,
 P_DPR, P_POOL0, P_DPERC, P_G2, P_G3, P_G4, P_GENK, P_TECHQ, P_TECHB,
 P_INF0, P_INFG, P_ECOF) = [P(r) for r in range(2, 23)]

# ---------------- 情景表 ----------------
HEADERS = ["回合", "训练%", "推理%", "研发%\n(自动)", "电力\n供给", "数据中心\n容量", "代际\n系数",
           "理论\n算力", "电力上限\n算力", "实际\n算力", "电力\n受限", "训练\n算力", "数据\n产出",
           "可用\n数据", "训练需\n数据", "数据\n满足率", "有效\n训练", "累积\n训练", "代际",
           "数据池\n余量", "研发\n算力", "科技点", "科技\n等级", "推理\n算力", "推理\n需求",
           "推理\n供给率", "经济\n系数", "芯片\n产出", "芯片\n可用", "芯片\n投资",
           "新增容量\n(次回合)", "芯片\n库存", "军力\n指数"]
PCT = {"B", "C", "D", "P", "Z"}
DEC2 = {"G", "AA"}
SCENARIOS = [
    ("情景A均衡", "均衡流：三通道大致均分", 0.34, 0.33),
    ("情景B奇点", "奇点冲刺流：算力压倒性投入训练", 0.65, 0.20),
    ("情景C军备", "军备优先流：推理喂饱军队与经济", 0.15, 0.65),
    ("情景D科研", "科研经济流：研发滚雪球", 0.20, 0.30),
]
T0, TURNS = 4, 20  # 数据起始行 / 回合数

for name, desc, tr, inf in SCENARIOS:
    s = wb.create_sheet(name)
    t = s.cell(row=1, column=1, value=f"{name} —— {desc}（训练 {tr:.0%} / 推理 {inf:.0%} / 研发 {1-tr-inf:.0%}）")
    t.font = Font(name=ARIAL, bold=True, size=13)
    s.cell(row=2, column=1, value="蓝色列逐回合可调；其余全部为公式").font = Font(name=ARIAL, size=9, color="808080")
    for i, h in enumerate(HEADERS, 1):
        c = s.cell(row=3, column=i, value=h)
        c.font, c.fill, c.alignment = HDR_FONT, HDR_FILL, WRAP_CENTER
    s.row_dimensions[3].height = 30
    s.column_dimensions["A"].width = 5
    for i in range(2, 34):
        s.column_dimensions[get_column_letter(i)].width = 8.5
    s.freeze_panes = "B4"

    for t_i in range(1, TURNS + 1):
        r = T0 + t_i - 1
        p = r - 1
        first = t_i == 1
        F = {}
        F["A"] = t_i
        F["B"] = tr
        F["C"] = inf
        F["D"] = f"=1-B{r}-C{r}"
        F["E"] = f"={P_POW0}+{P_POWG}*(A{r}-1)"
        F["F"] = f"={P_DC0}" if first else f"=F{p}+AE{p}"
        F["G"] = "=1" if first else f"=1+{P_GENK}*(S{p}-1)"
        F["H"] = f"=F{r}*G{r}"
        F["I"] = f"=E{r}/{P_E_PER}"
        F["J"] = f"=MIN(H{r},I{r})"
        F["K"] = f'=IF(I{r}<H{r},"⚡受限","")'
        F["L"] = f"=J{r}*B{r}"
        F["M"] = (f"={P_REG}*{P_DPR}" if first
                  else f"={P_REG}*{P_DPR}*(1+{P_TECHB}*W{p})*AA{p}")
        F["N"] = f"={P_POOL0}+M{r}" if first else f"=T{p}+M{r}"
        F["O"] = f"=L{r}*{P_DPERC}"
        F["P"] = f"=IF(O{r}=0,1,MIN(1,N{r}/O{r}))"
        F["Q"] = f"=L{r}*P{r}"
        F["R"] = f"=Q{r}" if first else f"=R{p}+Q{r}"
        F["S"] = f"=1+IF(R{r}>={P_G2},1,0)+IF(R{r}>={P_G3},1,0)+IF(R{r}>={P_G4},1,0)"
        F["T"] = f"=N{r}-O{r}*P{r}"
        F["U"] = f"=J{r}*D{r}"
        F["V"] = f"=U{r}" if first else f"=V{p}+U{r}"
        F["W"] = f"=INT(V{r}/{P_TECHQ})"
        F["X"] = f"=J{r}*C{r}"
        F["Y"] = f"={P_INF0}+{P_INFG}*(A{r}-1)"
        F["Z"] = f"=MIN(1,X{r}/Y{r})"
        F["AA"] = f"={P_ECOF}+(1-{P_ECOF})*Z{r}"
        F["AB"] = f"={P_FAB}" if first else f"={P_FAB}*(1+{P_TECHB}*W{p})*AA{p}"
        F["AC"] = f"={P_CHIP0}+AB{r}" if first else f"=AF{p}+AB{r}"
        F["AD"] = f"=MIN(AC{r},{P_INVCAP})"
        F["AE"] = f"=AD{r}/{P_CAPCOST}"
        F["AF"] = f"=AC{r}-AD{r}"
        F["AG"] = f"=X{r}*G{r}"
        for col, val in F.items():
            c = s[f"{col}{r}"]
            c.value = val
            c.font = BLUE if col in ("B", "C") else BLACK
            c.border = THIN
            if col in PCT:
                c.number_format = "0%"
            elif col in DEC2:
                c.number_format = "0.00"
            elif col != "K":
                c.number_format = "0"

# ---------------- 对比 ----------------
cmp_ = wb.create_sheet("对比")
cmp_.cell(row=1, column=1, value="四种分配策略 · 20 回合对比").font = Font(name=ARIAL, bold=True, size=14)
sum_headers = ["情景", "训练%", "推理%", "研发%", "期末\n代际", "Gen2\n达成回合", "Gen3\n达成回合",
               "期末\n实际算力", "期末\n军力指数", "期末\n科技等级", "期末\n经济系数", "数据满足率\n(末5回合)"]
for i, h in enumerate(sum_headers, 1):
    c = cmp_.cell(row=3, column=i, value=h)
    c.font, c.fill, c.alignment = HDR_FONT, HDR_FILL, WRAP_CENTER
cmp_.row_dimensions[3].height = 30
cmp_.column_dimensions["A"].width = 14
for i in range(2, 13):
    cmp_.column_dimensions[get_column_letter(i)].width = 11

LAST = T0 + TURNS - 1
for i, (name, desc, tr, inf) in enumerate(SCENARIOS):
    r = 4 + i
    q = f"'{name}'"
    cells = [
        (1, name.replace("情景", "") , BLACK, None),
        (2, f"={q}!B{T0}", GREEN, "0%"),
        (3, f"={q}!C{T0}", GREEN, "0%"),
        (4, f"={q}!D{T0}", GREEN, "0%"),
        (5, f"={q}!S{LAST}", GREEN, "0"),
        (6, f'=IFERROR(MATCH(2,{q}!S{T0}:S{LAST},0),"—")', GREEN, "0"),
        (7, f'=IFERROR(MATCH(3,{q}!S{T0}:S{LAST},0),"—")', GREEN, "0"),
        (8, f"={q}!J{LAST}", GREEN, "0"),
        (9, f"={q}!AG{LAST}", GREEN, "0"),
        (10, f"={q}!W{LAST}", GREEN, "0"),
        (11, f"={q}!AA{LAST}", GREEN, "0.00"),
        (12, f"=AVERAGE({q}!P{LAST-4}:P{LAST})", GREEN, "0%"),
    ]
    for col, val, font, fmt in cells:
        c = cmp_.cell(row=r, column=col, value=val)
        c.font, c.border = font, THIN
        if fmt:
            c.number_format = fmt

def line_chart(title, col, anchor, y_title):
    ch = LineChart()
    ch.title, ch.height, ch.width = title, 8, 17
    ch.y_axis.title, ch.x_axis.title = y_title, "回合"
    for name, *_ in SCENARIOS:
        ws_ = wb[name]
        ser = Series(Reference(ws_, min_col=col, min_row=T0, max_row=LAST), title=name[3:])
        ch.series.append(ser)
    ch.set_categories(Reference(wb[SCENARIOS[0][0]], min_col=1, min_row=T0, max_row=LAST))
    cmp_.add_chart(ch, anchor)

line_chart("军力指数走势", 33, "A10", "军力指数")
line_chart("累积训练进度（代际竞赛）", 18, "J10", "累积训练")
line_chart("实际算力（注意电力墙）", 10, "A27", "实际算力")
line_chart("经济系数（推理饥饿的代价）", 27, "J27", "经济系数")

out = "/Users/jianmingxie/Documents/futrue conquer/M0_数值原型.xlsx"
wb.save(out)
print("saved", out)
